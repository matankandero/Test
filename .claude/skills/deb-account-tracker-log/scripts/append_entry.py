#!/usr/bin/env python3
"""
Append a new dated paragraph to an account's existing Activity cell in one of
the Selution account activity trackers, without inserting or reordering rows.

Never call ws.insert_rows() / ws.delete_rows() here — the tracker uses merged
cells for the Account/Owner columns, and row insertion does not reliably move
merged ranges in openpyxl (this has silently corrupted unrelated rows during
testing). This script only ever mutates the text value of one existing cell.

Usage:
    python append_entry.py --file "<path to tracker xlsx>" \
        --account "Soroka Cardio" --entry "<translated English text>" \
        --date "7.7" [--dry-run]

Run with --dry-run first, always. Only re-run without it after the user
has confirmed the printed old_text -> new_text diff.
"""

import argparse
import sys

try:
    import openpyxl
except ImportError:
    sys.exit(
        "openpyxl is required (pip install openpyxl). This script must run "
        "locally where the tracker files actually live, not in a sandbox."
    )


def find_header(ws):
    """Locate the real Account/Account Manager header row (not the small,
    lowercase-typo'd 'Planned Actions' mini-table header elsewhere in the
    sheet) and return (header_row, account_col, owner_col)."""
    for row in ws.iter_rows():
        values = {cell.column: (cell.value.strip() if isinstance(cell.value, str) else cell.value) for cell in row}
        account_col = next((col for col, v in values.items() if v == "Account"), None)
        owner_col = next((col for col, v in values.items() if v == "Account Manager"), None)
        if account_col is not None and owner_col is not None:
            return row[0].row, account_col, owner_col
    raise RuntimeError(
        "Could not find a header row containing 'Account' and 'Account Manager' cells. "
        "The sheet layout may have changed — stop and ask Matan rather than guessing."
    )


def merged_range_containing(ws, row, col):
    """Return the merged cell range that contains (row, col), or None."""
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col):
            return merged_range
    return None


def list_known_accounts(ws, header_row, account_col):
    names = []
    for row in ws.iter_rows(min_row=header_row + 1, min_col=account_col, max_col=account_col):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                names.append(cell.value.strip())
    return names


def find_account_block(ws, header_row, account_col, account_name):
    """Find the merged-cell block for account_name in the Account column.
    Returns (min_row, max_row) of the block. Raises with a helpful message
    (listing known accounts) if not found or ambiguous."""
    target = account_name.strip().lower()
    matches = []
    for row in ws.iter_rows(min_row=header_row + 1, min_col=account_col, max_col=account_col):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                if cell.value.strip().lower() == target:
                    matches.append(cell)

    if not matches:
        # fall back to substring match
        for row in ws.iter_rows(min_row=header_row + 1, min_col=account_col, max_col=account_col):
            for cell in row:
                if isinstance(cell.value, str) and target in cell.value.strip().lower():
                    matches.append(cell)

    if not matches:
        known = list_known_accounts(ws, header_row, account_col)
        raise RuntimeError(
            f"No account matching '{account_name}' found. Known accounts in this file: "
            + ", ".join(known)
        )
    if len(matches) > 1:
        raise RuntimeError(
            f"'{account_name}' matched multiple rows ambiguously: "
            + ", ".join(f"{c.value!r} (row {c.row})" for c in matches)
            + ". Stop and ask Matan which one this belongs to."
        )

    cell = matches[0]
    merged_range = merged_range_containing(ws, cell.row, cell.column)
    if merged_range is not None:
        return merged_range.min_row, merged_range.max_row
    return cell.row, cell.row


def find_activity_cell(ws, block_min_row, block_max_row, account_col, owner_col):
    """The Activity column is the one immediately after the Account Manager
    column, per the file's observed layout (Account, Account Manager,
    Activity, ...). Within the account's block, use the last row that
    already has non-empty text in that column — new entries get appended
    to the most recent existing paragraph cell, never to a blank row."""
    activity_col = owner_col + 1
    for row_idx in range(block_max_row, block_min_row - 1, -1):
        cell = ws.cell(row=row_idx, column=activity_col)
        if isinstance(cell.value, str) and cell.value.strip():
            return cell
    # nothing has text yet (shouldn't normally happen) — use the block's first row
    return ws.cell(row=block_min_row, column=activity_col)


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--file", required=True, help="Full path to the tracker .xlsx file")
    parser.add_argument("--account", required=True, help="Account name as it appears in the tracker (e.g. 'Soroka Cardio')")
    parser.add_argument("--entry", required=True, help="Translated English entry text to append")
    parser.add_argument("--date", required=True, help="Date of the meeting, in the file's D.M convention (e.g. '7.7')")
    parser.add_argument("--dry-run", action="store_true", help="Preview the change without writing the file")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.file)
    ws = wb.worksheets[0]

    header_row, account_col, owner_col = find_header(ws)
    block_min_row, block_max_row = find_account_block(ws, header_row, account_col, args.account)
    cell = find_activity_cell(ws, block_min_row, block_max_row, account_col, owner_col)

    old_text = cell.value or ""
    new_paragraph = f"{args.date} Update: {args.entry.strip()}"
    new_text = (old_text.rstrip() + "\n\n" + new_paragraph) if old_text.strip() else new_paragraph

    print(f"Account: {args.account}  (block rows {block_min_row}-{block_max_row}, writing to row {cell.row}, col {cell.column})")
    print("--- OLD (last 300 chars) ---")
    print(old_text[-300:])
    print("--- NEW (last 300 chars) ---")
    print(new_text[-300:])

    if args.dry_run:
        print("\n[dry run] No changes written.")
        return

    cell.value = new_text
    wb.save(args.file)
    print(f"\nSaved: {args.file}")


if __name__ == "__main__":
    main()
